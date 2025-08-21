from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from typing import List, Tuple, Dict, Any, Optional
import re
from difflib import SequenceMatcher

def _normalize_header(s: str) -> str:
    """Normalize header text for matching: remove punctuation, collapse whitespace, lowercase."""
    if s is None:
        return ""
    # replace non-word with space, collapse, lowercase
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", s)).strip().lower()

def find_header_row_and_cols(sheet) -> Dict[str, Tuple[int,int]]:
    """
    DEPRECATED: returns first textual occurrences map. Kept for compatibility.
    Use locate_header_row() + headers_from_row() instead.
    """
    headers = {}
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val and isinstance(val, str) and val.strip():
                key = val.strip().lower()
                if key not in headers:
                    headers[key] = (r, c)
    return headers

def locate_header_row(sheet, key_header_norm: str, top_rows_prefer: int = 20) -> Tuple[int,int]:
    """
    Find best header row/col for the given normalized key header text.
    Strategy:
      - Prefer an occurrence within the top_rows_prefer rows (typical headers).
      - If multiple in top rows choose the one with most textual cells in that row.
      - Otherwise search entire sheet for the best row (most textual cells containing the key).
      - Fallback to first exact match anywhere.
    Returns (row, col).
    """
    max_col = sheet.max_column
    max_row = sheet.max_row
    candidates = []

    # gather exact matches (case-insensitive & normalized)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = sheet.cell(row=r, column=c).value
            if v and isinstance(v, str) and _normalize_header(v) == key_header_norm:
                # count textual cells in the row to prefer header-like rows
                cnt = 0
                for cc in range(1, max_col + 1):
                    vv = sheet.cell(row=r, column=cc).value
                    if vv and isinstance(vv, str) and vv.strip():
                        cnt += 1
                candidates.append((r, c, cnt))

    if not candidates:
        raise ValueError(f"Key header '{key_header_norm}' not found in sheet")

    # prefer occurrences within top_rows_prefer
    top_candidates = [t for t in candidates if t[0] <= top_rows_prefer]
    if top_candidates:
        # choose the one with highest textual count, tie-breaker: smallest row
        top_candidates.sort(key=lambda x: (-x[2], x[0], x[1]))
        return (top_candidates[0][0], top_candidates[0][1])

    # otherwise choose candidate with largest textual count, tie-breaker smallest row
    candidates.sort(key=lambda x: (-x[2], x[0], x[1]))
    return (candidates[0][0], candidates[0][1])

def headers_from_row(sheet, header_row: int) -> Dict[str, Tuple[int,int,str]]:
    """
    Return mapping normalized_header -> (row, col, display_value) for cells only in header_row.
    This prevents picking up data cells as headers and uses normalized keys for matching.
    Coerce non-str header values to str so numeric headers are preserved.
    """
    hdrs: Dict[str, Tuple[int,int,str]] = {}
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=header_row, column=c).value
        if v is not None:
            display = str(v).strip()
            if display:
                norm = _normalize_header(display)
                if norm:  # only keep if normalization non-empty
                    # keep the first occurrence for this normalized key (left-most)
                    if norm not in hdrs:
                        hdrs[norm] = (header_row, c, display)
    return hdrs

def build_row_dict(sheet, header_row: int, key_col: int) -> Dict[str, int]:
    """Build mapping from normalized key value -> row number (first occurrence after header_row)."""
    row_dict: Dict[str, int] = {}
    for r in range(header_row + 1, sheet.max_row + 1):
        key = sheet.cell(row=r, column=key_col).value
        if key not in (None, ""):
            row_dict[_normalize_header(str(key))] = r
    return row_dict

def _normalize_key(k: Any) -> str:
    if k is None:
        return ""
    return str(k).strip()

def _safe_eq(a: Any, b: Any) -> bool:
    """
    Return True when values are equal and not both empty.
    If both empty -> False (no highlight).
    """
    if a in (None, "") and b in (None, ""):
        return False
    return _normalize_key(a) == _normalize_key(b)

def _sort_keys(keys: List[str]) -> List[str]:
    # Try numeric sort where possible, fallback to string sort
    def keyfn(x: str):
        try:
            return (0, float(x))
        except Exception:
            return (1, x)
    return sorted(keys, key=keyfn)

def compare_excel_files(file_list: List[str], options: Optional[Dict[str, Any]] = None, **kwargs) -> Any:
    """
    Generalized entry for Excel comparison.
    - file_list: [original_path, new_path]
    - options: dict of optional keys (key_header, compare_mode, weights, sample_size,
      top_rows_prefer, match_color, add_color, del_color, output_path, progress_cb, return_meta)
    - kwargs are merged into options for backward compatibility.

    Returns: Workbook (if not saved) or output_path (if saved) or (result, meta) when return_meta True.
    """
    # merge options + kwargs (backwards compatible)
    opts: Dict[str, Any] = {}
    if options and isinstance(options, dict):
        opts.update(options)
    opts.update(kwargs)

    # defaults and option extraction
    key_header = opts.get("key_header", "S.no")
    compare_mode = opts.get("compare_mode", "by_key")
    match_color = opts.get("match_color", "FFFF00")
    add_color = opts.get("add_color", "C6EFCE")
    del_color = opts.get("del_color", "FFC7CE")
    weight_name = float(opts.get("weight_name", 0.25))
    weight_data = float(opts.get("weight_data", 0.75))
    min_score = float(opts.get("min_score", 0.20))
    sample_size = int(opts.get("sample_size", 40))
    top_rows_prefer = int(opts.get("top_rows_prefer", 20))
    output_path_opt = opts.get("output_path")
    progress_cb = opts.get("progress_cb")  # callable(percent:int)
    return_meta = bool(opts.get("return_meta", False))

    if len(file_list) < 2:
        raise ValueError("At least two files are required for comparison.")

    def _progress(p: int):
        try:
            if callable(progress_cb):
                progress_cb(int(p))
        except Exception:
            pass

    _progress(1)
    wb1 = load_workbook(file_list[0], data_only=True)
    wb2 = load_workbook(file_list[1], data_only=True)
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]

    key_header_norm = _normalize_header(key_header)

    # locate header rows (prefer top rows)
    key_row1, key_col1 = locate_header_row(sheet1, key_header_norm, top_rows_prefer=top_rows_prefer)
    key_row2, key_col2 = locate_header_row(sheet2, key_header_norm, top_rows_prefer=top_rows_prefer)

    _progress(5)
    # build row dictionaries mapping normalized key value -> source row
    row_dict1 = build_row_dict(sheet1, key_row1, key_col1)
    row_dict2 = build_row_dict(sheet2, key_row2, key_col2)

    # extract headers only from the header rows (normalized keys)
    headers1_row = headers_from_row(sheet1, key_row1)
    headers2_row = headers_from_row(sheet2, key_row2)

    # try to match headers by data+name first (use provided tuning params)
    col_match = match_columns_by_data(headers1_row, headers2_row, sheet1, sheet2, key_row1, key_row2,
                                      weight_name=weight_name, weight_data=weight_data,
                                      min_score=min_score, sample_size=sample_size)

    # Build ordered header list (preserve file1 left-to-right, then unmatched file2 headers)
    def _ordered_norms_from_map(hmap: Dict[str, Tuple[int,int,str]]) -> List[str]:
        items = [(norm, t[1]) for norm, t in hmap.items()]  # (norm, col)
        return [norm for norm, _col in sorted(items, key=lambda x: x[1])]

    hdrs1_ordered = _ordered_norms_from_map(headers1_row)
    hdrs2_ordered = _ordered_norms_from_map(headers2_row)

    header_list: List[str] = []
    # ensure key header first if available
    if key_header_norm in headers1_row or key_header_norm in headers2_row:
        header_list.append(key_header_norm)

    used_in_file2 = set()
    for h in hdrs1_ordered:
        if h == key_header_norm:
            continue
        if h not in header_list:
            header_list.append(h)
        partner = col_match.get(h)
        if partner:
            used_in_file2.add(partner)

    for h in hdrs2_ordered:
        if h == key_header_norm:
            continue
        if h in header_list or h in used_in_file2:
            continue
        header_list.append(h)

    if not header_list:
        raise ValueError("No headers found in either file to compare.")

    _progress(15)
    # --- create output workbook/worksheet and styles ---
    result_wb = Workbook()
    ws = result_wb.active
    ws.title = "Comparison"

    match_fill = PatternFill(start_color=match_color, end_color=match_color, fill_type="solid")
    add_fill = PatternFill(start_color=add_color, end_color=add_color, fill_type="solid")
    del_fill = PatternFill(start_color=del_color, end_color=del_color, fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.cell(row=1, column=1, value="Legend:")
    ws.cell(row=2, column=1, value="Yellow = Match")
    ws.cell(row=3, column=1, value="Red = Removed/Changed (File 1)")
    ws.cell(row=4, column=1, value="Green = Added/Changed (File 2)")

    ws.cell(row=6, column=1, value="Rows: File 1/File 2")

    header_out_col = 2
    header_map = []
    for hnorm in header_list:
        r1c = headers1_row.get(hnorm)
        if hnorm in headers1_row and hnorm in col_match:
            partner_key = col_match.get(hnorm)
            r2c = headers2_row.get(partner_key)
        else:
            r2c = headers2_row.get(hnorm)

        if r1c is None and r2c is not None:
            rev = None
            for k, v in col_match.items():
                if v == hnorm:
                    rev = k
                    break
            r1c = headers1_row.get(rev) if rev else None

        c1 = r1c[1] if r1c else None
        c2 = r2c[1] if r2c else None
        disp1 = r1c[2] if r1c else (r2c[2] if r2c else hnorm)
        disp2 = r2c[2] if r2c else (r1c[2] if r1c else hnorm)

        ws.cell(row=6, column=header_out_col, value=f"{disp1} (File 1)")
        ws.cell(row=6, column=header_out_col + 1, value=f"{disp2} (File 2)")
        ws.cell(row=6, column=header_out_col).font = Font(bold=True)
        ws.cell(row=6, column=header_out_col + 1).font = Font(bold=True)
        header_map.append((hnorm, header_out_col, header_out_col + 1, c1, c2, disp1, disp2))
        header_out_col += 2

    _progress(30)
    # build rows_to_compare according to mode
    rows_to_compare: List[Tuple[Optional[int], Optional[int], str]] = []
    if compare_mode == "by_key":
        all_keys = _sort_keys(list(set(row_dict1.keys()) | set(row_dict2.keys())))
        for key_norm in all_keys:
            r1 = row_dict1.get(key_norm)
            r2 = row_dict2.get(key_norm)
            # try to display the actual key value (from key column) rather than raw sheet row numbers
            disp1 = ""
            disp2 = ""
            try:
                if r1:
                    raw = sheet1.cell(row=r1, column=key_col1).value
                    disp1 = str(raw) if raw not in (None, "") else ""
            except Exception:
                disp1 = ""
            try:
                if r2:
                    raw = sheet2.cell(row=r2, column=key_col2).value
                    disp2 = str(raw) if raw not in (None, "") else ""
            except Exception:
                disp2 = ""
            label = f"{disp1 or (r1 or '')}/{disp2 or (r2 or '')}"
            rows_to_compare.append((r1, r2, label))
    elif compare_mode == "by_row":
        max_after = max(sheet1.max_row - key_row1, sheet2.max_row - key_row2)
        for offset in range(1, max_after + 1):
            r1 = key_row1 + offset if key_row1 + offset <= sheet1.max_row else None
            r2 = key_row2 + offset if key_row2 + offset <= sheet2.max_row else None
            # prefer showing key column values at these rows (user-friendly), fall back to sheet row numbers
            disp1 = ""
            disp2 = ""
            try:
                if r1 and key_col1:
                    raw = sheet1.cell(row=r1, column=key_col1).value
                    disp1 = str(raw) if raw not in (None, "") else ""
            except Exception:
                disp1 = ""
            try:
                if r2 and key_col2:
                    raw = sheet2.cell(row=r2, column=key_col2).value
                    disp2 = str(raw) if raw not in (None, "") else ""
            except Exception:
                disp2 = ""
            label = f"{disp1 or (r1 or '')}/{disp2 or (r2 or '')}"
            rows_to_compare.append((r1, r2, label))
    else:
        raise ValueError("compare_mode must be 'by_key' or 'by_row'")

    total = max(1, len(rows_to_compare))
    # write comparisons
    for i, (r1, r2, label) in enumerate(rows_to_compare):
        out_row = 7 + i
        ws.cell(row=out_row, column=1, value=label)
        if i % 2 == 1:
            for _, out_c1, out_c2, _, _, _, _ in header_map:
                ws.cell(row=out_row, column=out_c1).fill = alt_fill
                ws.cell(row=out_row, column=out_c2).fill = alt_fill

        for _, out_c1, out_c2, col1, col2, _, _ in header_map:
            val1 = sheet1.cell(row=r1, column=col1).value if (r1 and col1) else None
            val2 = sheet2.cell(row=r2, column=col2).value if (r2 and col2) else None

            cell1 = ws.cell(row=out_row, column=out_c1, value=val1)
            cell2 = ws.cell(row=out_row, column=out_c2, value=val2)

            if val1 in (None, "") and val2 in (None, ""):
                continue
            if _safe_eq(val1, val2):
                cell1.fill = match_fill
                cell2.fill = match_fill
            else:
                if val1 not in (None, ""):
                    cell1.fill = del_fill
                if val2 not in (None, ""):
                    cell2.fill = add_fill

        # progress update
        if (i % 20) == 0:
            _progress(30 + int(60 * (i / total)))

    _progress(95)
    # freeze panes, autosize, autofilter (unchanged)
    ws.freeze_panes = ws["B7"]
    for col in ws.columns:
        try:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(120, max_length + 2)
        except Exception:
            pass

    try:
        last_col_idx = header_out_col - 1
        last_col_letter = ws.cell(row=6, column=last_col_idx).column_letter
        ws.auto_filter.ref = f"A6:{last_col_letter}6"
    except Exception:
        ws.auto_filter.ref = ws.dimensions

    _progress(100)
    # save if requested
    if output_path_opt:
        save_comparison_result(result_wb, output_path_opt)
        result = output_path_opt
    else:
        result = result_wb

    meta = {"rows_compared": len(rows_to_compare), "key_header": key_header, "compare_mode": compare_mode}
    if return_meta:
        return result, meta
    return result

def save_comparison_result(result_wb: Workbook, output_path: str) -> None:
    result_wb.save(output_path)

def _sample_column_values(sheet, header_row: int, col: int, max_samples: int = 50) -> List[str]:
    vals = []
    for r in range(header_row + 1, min(sheet.max_row + 1, header_row + 1 + max_samples*5)):
        v = sheet.cell(row=r, column=col).value
        if v not in (None, ""):
            vals.append(_normalize_key(v))
            if len(vals) >= max_samples:
                break
    return vals

def _jaccard(a: List[str], b: List[str]) -> float:
    sa = set(a)
    sb = set(b)
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    inter = sa & sb
    union = sa | sb
    return len(inter) / len(union)

def _name_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def match_columns_by_data(headers1_row: Dict[str, Tuple[int,int,str]],
                          headers2_row: Dict[str, Tuple[int,int,str]],
                          sheet1, sheet2,
                          header_row1: int, header_row2: int,
                          weight_name: float = 0.3,
                          weight_data: float = 0.7,
                          min_score: float = 0.25,
                          sample_size: int = 40) -> Dict[str, str]:
    """
    Return mapping norm_header1 -> norm_header2 for matched columns.
    Greedy match using combined name + data similarity.
    """
    scores = []
    # pre-sample data for each column
    samples1 = {h: _sample_column_values(sheet1, header_row1, c, sample_size) for h, (r,c,disp) in headers1_row.items()}
    samples2 = {h: _sample_column_values(sheet2, header_row2, c, sample_size) for h, (r,c,disp) in headers2_row.items()}

    for h1, (r1, c1, d1) in headers1_row.items():
        for h2, (r2, c2, d2) in headers2_row.items():
            name_s = _name_ratio(d1, d2)
            data_s = _jaccard(samples1.get(h1, []), samples2.get(h2, []))
            score = weight_name * name_s + weight_data * data_s
            scores.append((score, h1, h2, name_s, data_s))

    # sort descending
    scores.sort(reverse=True, key=lambda x: x[0])

    mapped1 = set()
    mapped2 = set()
    mapping = {}

    for score, h1, h2, name_s, data_s in scores:
        if score < min_score:
            break
        if h1 in mapped1 or h2 in mapped2:
            continue
        mapping[h1] = h2
        mapped1.add(h1)
        mapped2.add(h2)

    return mapping

if __name__ == "__main__":
    f1 = r"C:\Users\112957\OneDrive - UL Solutions\Desktop\Automation projects\Difference Checker\test_samples\NS003_HW_3_A.xlsx"
    f2 = r"C:\Users\112957\OneDrive - UL Solutions\Desktop\Automation projects\Difference Checker\test_samples\NS003_HW_4.0_A.xlsx"
    wb = compare_excel_files([f1, f2], key_header="S.no", compare_mode="by_row")
    save_comparison_result(wb, r"C:\Users\112957\OneDrive - UL Solutions\Desktop\Automation projects\Difference Checker\test_samples\excel_comparison_result.xlsx")